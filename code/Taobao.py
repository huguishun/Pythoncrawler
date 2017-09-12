import requests
from openpyxl import Workbook
import re
import time
import os
def File():
	os.makedirs(os.path.join("F:\Taobao", pagename))
	os.makedirs(os.path.join("F:\Taobao", page_img))
	get_url()
def get_url():
	start_url = "https://s.taobao.com/search?q="+keyword
	user_agent = "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:35.0) Gecko/20100101 Firefox/35.0"
	headers = {"User-Agent":user_agent}
	res = requests.get(start_url,headers)
	wb = Workbook()
	for a in range(page):
		time.sleep(10)
		url_page = "https://s.taobao.com/search?q="+keyword+"&imgfile=&commend=all&ssid=s5-e&search_type=item&sourceId=tb.index&spm=a21bo.50862.201856-taobao-item.1&ie=utf8&bcoffset=4&ntoffset=4&p4ppushleft=1%2C48&s="+str(a*44)
		url_data = requests.get(url_page)
		# 价格
		plt = re.findall(r'\"view_price\"\:\"[\d\.]*\"',url_data.text)
		# 商品名
		tlt = re.findall(r'\"raw_title\"\:\".*?\"', url_data.text)
		# 店铺名
		nic = re.findall(r'\"nick\"\:\".*?\"',url_data.text)
		# 商品图片
		img = re.findall(r'\"pic_url\"\:\".*?\"',url_data.text)
		if a >= 1:
			ws = wb.create_sheet()
		for i in range(len(plt)):
			ws = wb.active
			ws = wb.worksheets[a]
			price = eval(plt[i].split(':')[1])
			title = eval(tlt[i].split(':')[1])
			nick = eval(nic[i].split(':')[1])
			img_url = eval(img[i].split(':')[1])
			all_url = requests.get('http:'+img_url)
			os.chdir("F:\Taobao\\"+page_img)
			title_txt = re.sub('[\r\n\t\*\?\<\>\|\:\/ ]', '', title)
			f = open(title_txt+'.jpg', 'wb')
			f.write(all_url.content) 
			f.close()
			os.chdir("F:\Taobao\\"+pagename)
			ws['A1'].value = '价格'
			ws['B1'].value = '商品名'
			ws['C1'].value = '店铺名'
			ws['D1'].value = "图片URL"
			i+=1
			ws['A%d'%(i+1)].value = price
			ws['B%d'%(i+1)].value = title
			ws['C%d'%(i+1)].value = nick
			ws['D%d'%(i+1)].value = img_url
			print('正在获取第'+str(a+1)+'页，第'+str(i)+'个商品信息')
	filename = input("请输入文件名：")+'.xlsx'
	wb.save(filename)
	print("任务完成，数据已存入。")
if __name__ == '__main__':
	keyword = str(input("请输入需求获取的商品名："))
	page = int(input("请输入商品页数："))
	pagename = input("请为存储信息的文件夹命名:")
	page_img = input("请为存储图片信息的文件夹命名：")
	File()
